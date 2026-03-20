"""
============================================================
00_build_ecrf_specs.py
Purpose: Generate eCRF Configuration Specification Workbook
Study:   XYZ-2026 Phase II Diabetes Trial
Author:  Careen Chowrappa
Date:    March 2026
============================================================
Simulates the work a Configuration & QC Specialist does when
setting up a new clinical trial database in an EDC system.
Maps to ICON JD: "Configuration and setup of clinical trial
databases and electronic data capture (EDC) systems"
============================================================
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("=" * 60)
print("eCRF SPECIFICATION BUILDER — Study XYZ-2026")
print("=" * 60)

# ============================================================
# DOMAIN SPECIFICATIONS
# ============================================================

# DM - Demographics
dm_spec = pd.DataFrame([
    ['DM-001', 'USUBJID', 'Subject Identifier', 'Text', 15, 'Yes', '—',
     'Format: XYZ-2026-SSS-PPPP', 'USUBJID', 'Unique subject ID across all sites'],
    ['DM-002', 'SITEID', 'Site Number', 'Coded', 3, 'Yes', '001|002|003|004|005',
     'Must be valid site code', 'SITEID', 'Investigator site identifier'],
    ['DM-003', 'RFICDTC', 'Informed Consent Date', 'Date', 10, 'Yes', '—',
     'Must be ≤ today; format YYYY-MM-DD', 'RFICDTC', 'Date subject signed ICF'],
    ['DM-004', 'BRTHDTC', 'Date of Birth', 'Date', 10, 'Yes', '—',
     'Subject must be 18-75 at consent', 'BRTHDTC', 'Used to derive AGE'],
    ['DM-005', 'AGE', 'Age at Consent', 'Integer', 3, 'Yes', '—',
     'Must be 18-75; must match DOB calculation', 'AGE', 'Derived from BRTHDTC and RFICDTC'],
    ['DM-006', 'SEX', 'Sex', 'Coded', 1, 'Yes', 'M|F',
     'Must be M or F', 'SEX', 'Biological sex at birth'],
    ['DM-007', 'RACE', 'Race', 'Coded', 40, 'Yes',
     'WHITE|BLACK OR AFRICAN AMERICAN|ASIAN|AMERICAN INDIAN OR ALASKA NATIVE|OTHER',
     'At least one must be selected', 'RACE', 'Per FDA demographic guidance'],
    ['DM-008', 'ETHNIC', 'Ethnicity', 'Coded', 30, 'Yes',
     'HISPANIC OR LATINO|NOT HISPANIC OR LATINO',
     'Must select one', 'ETHNIC', 'Per FDA demographic guidance'],
    ['DM-009', 'ARM', 'Treatment Arm', 'Coded', 20, 'Yes',
     'DRUG A 10MG|DRUG B 25MG|PLACEBO',
     'Assigned by IXRS at randomization', 'ARM', 'Randomized treatment group'],
    ['DM-010', 'RFSTDTC', 'Date of First Study Drug', 'Date', 10, 'Yes', '—',
     'Must be ≥ RFICDTC', 'RFSTDTC', 'First dose of study treatment'],
    ['DM-011', 'RFENDTC', 'Date of Last Study Drug', 'Date', 10, 'Cond', '—',
     'Must be ≥ RFSTDTC; required at EOS', 'RFENDTC', 'Last dose date'],
    ['DM-012', 'DTHFL', 'Death Flag', 'Coded', 1, 'Cond', 'Y|N',
     'If Y, death date required', 'DTHFL', 'Subject death indicator'],
], columns=['Field_ID', 'Variable', 'Label', 'DataType', 'Length',
           'Required', 'CodedList', 'ValidationRule', 'SDTM_Map', 'Description'])

# AE - Adverse Events
ae_spec = pd.DataFrame([
    ['AE-001', 'USUBJID', 'Subject Identifier', 'Text', 15, 'Yes', '—',
     'Must exist in DM domain', 'USUBJID', 'Link to demographics'],
    ['AE-002', 'AESEQ', 'AE Sequence Number', 'Integer', 3, 'Auto', '—',
     'Auto-incremented per subject', 'AESEQ', 'Unique within subject'],
    ['AE-003', 'AETERM', 'Reported Term', 'Text', 200, 'Yes', '—',
     'Free text verbatim from investigator', 'AETERM', 'Verbatim AE as reported'],
    ['AE-004', 'AEDECOD', 'Dictionary-Derived Term', 'Text', 200, 'Yes', 'MedDRA v27.0',
     'Must be coded if AETERM present', 'AEDECOD', 'MedDRA preferred term'],
    ['AE-005', 'AEBODSYS', 'Body System', 'Text', 200, 'Auto', 'MedDRA SOC',
     'Auto-populated from MedDRA coding', 'AEBODSYS', 'System Organ Class'],
    ['AE-006', 'AESTDTC', 'Start Date', 'Date', 10, 'Yes', '—',
     'Must be ≥ RFICDTC (consent); format YYYY-MM-DD', 'AESTDTC', 'AE onset date'],
    ['AE-007', 'AEENDTC', 'End Date', 'Date', 10, 'Cond', '—',
     'Must be ≥ AESTDTC; required if AEOUT=RECOVERED', 'AEENDTC', 'AE resolution date'],
    ['AE-008', 'AESEV', 'Severity', 'Coded', 10, 'Yes', 'MILD|MODERATE|SEVERE',
     'CTCAE grading; if SEVERE check AESER', 'AESEV', 'AE severity grade'],
    ['AE-009', 'AESER', 'Serious', 'Coded', 1, 'Yes', 'Y|N',
     'If Y, SAE form required within 24hrs', 'AESER', 'Meets ICH E2A serious criteria'],
    ['AE-010', 'AEREL', 'Causality', 'Coded', 20, 'Yes',
     'NOT RELATED|UNLIKELY|POSSIBLE|PROBABLE|DEFINITE',
     'Investigator assessment required', 'AEREL', 'Relationship to study drug'],
    ['AE-011', 'AEOUT', 'Outcome', 'Coded', 20, 'Yes',
     'RECOVERED/RESOLVED|RECOVERING/RESOLVING|NOT RECOVERED/NOT RESOLVED|'
     'RECOVERED WITH SEQUELAE|FATAL|UNKNOWN',
     'If FATAL, death date required in DM', 'AEOUT', 'AE outcome at last assessment'],
    ['AE-012', 'AEACN', 'Action Taken', 'Coded', 30, 'Yes',
     'NONE|DOSE REDUCED|DRUG INTERRUPTED|DRUG WITHDRAWN|NOT APPLICABLE',
     'Must be documented for all AEs', 'AEACN', 'Action taken with study drug'],
], columns=['Field_ID', 'Variable', 'Label', 'DataType', 'Length',
           'Required', 'CodedList', 'ValidationRule', 'SDTM_Map', 'Description'])

# VS - Vital Signs
vs_spec = pd.DataFrame([
    ['VS-001', 'USUBJID', 'Subject Identifier', 'Text', 15, 'Yes', '—',
     'Must exist in DM domain', 'USUBJID', 'Link to demographics'],
    ['VS-002', 'VISITNUM', 'Visit Number', 'Integer', 3, 'Yes', '1|2|3|4|5|6',
     'Per visit schedule in protocol', 'VISITNUM', 'Protocol visit number'],
    ['VS-003', 'VISIT', 'Visit Name', 'Coded', 30, 'Yes',
     'SCREENING|BASELINE|WEEK 4|WEEK 8|WEEK 12|END OF STUDY',
     'Must match VISITNUM', 'VISIT', 'Protocol visit label'],
    ['VS-004', 'VSDTC', 'Assessment Date', 'Date', 10, 'Yes', '—',
     'Must be within visit window (±3 days)', 'VSDTC', 'Date vitals were taken'],
    ['VS-005', 'SYSBP', 'Systolic Blood Pressure', 'Integer', 3, 'Yes', '—',
     'Range: 70-250 mmHg; must be > DIABP', 'VSSTRESN (VSTESTCD=SYSBP)', 'Systolic BP in mmHg'],
    ['VS-006', 'DIABP', 'Diastolic Blood Pressure', 'Integer', 3, 'Yes', '—',
     'Range: 40-150 mmHg; must be < SYSBP', 'VSSTRESN (VSTESTCD=DIABP)', 'Diastolic BP in mmHg'],
    ['VS-007', 'PULSE', 'Heart Rate', 'Integer', 3, 'Yes', '—',
     'Range: 40-180 bpm', 'VSSTRESN (VSTESTCD=PULSE)', 'Beats per minute'],
    ['VS-008', 'TEMP', 'Temperature', 'Decimal', 4, 'Yes', '—',
     'Range: 35.0-40.0 °C', 'VSSTRESN (VSTESTCD=TEMP)', 'Degrees Celsius'],
    ['VS-009', 'WEIGHT', 'Weight', 'Decimal', 5, 'Yes', '—',
     'Range: 30-300 kg; flag if >10% change from baseline', 'VSSTRESN (VSTESTCD=WEIGHT)', 'Kilograms'],
    ['VS-010', 'HEIGHT', 'Height', 'Decimal', 5, 'Cond', '—',
     'Collected at Screening only; range: 100-220 cm', 'VSSTRESN (VSTESTCD=HEIGHT)', 'Centimeters; screening only'],
], columns=['Field_ID', 'Variable', 'Label', 'DataType', 'Length',
           'Required', 'CodedList', 'ValidationRule', 'SDTM_Map', 'Description'])

# LB - Laboratory
lb_spec = pd.DataFrame([
    ['LB-001', 'USUBJID', 'Subject Identifier', 'Text', 15, 'Yes', '—',
     'Must exist in DM domain', 'USUBJID', 'Link to demographics'],
    ['LB-002', 'VISITNUM', 'Visit Number', 'Integer', 3, 'Yes', '—',
     'Per visit schedule', 'VISITNUM', 'Protocol visit number'],
    ['LB-003', 'LBDTC', 'Collection Date', 'Date', 10, 'Yes', '—',
     'Must be within visit window', 'LBDTC', 'Sample collection date'],
    ['LB-004', 'LBTESTCD', 'Lab Test Code', 'Coded', 8, 'Yes',
     'GLUC|HBA1C|ALT|AST|CREAT|BUN|CHOL|TRIG|HDL|LDL',
     'Standard LOINC-mapped test codes', 'LBTESTCD', 'Short name for lab test'],
    ['LB-005', 'LBTEST', 'Lab Test Name', 'Text', 40, 'Auto', '—',
     'Auto-derived from LBTESTCD', 'LBTEST', 'Full descriptive name'],
    ['LB-006', 'LBORRES', 'Result (Original)', 'Text', 20, 'Yes', '—',
     'As reported by central lab', 'LBORRES', 'Result in original units'],
    ['LB-007', 'LBORRESU', 'Original Units', 'Coded', 20, 'Yes',
     'mg/dL|%|U/L|umol/L|mmol/L',
     'Must match test-specific unit', 'LBORRESU', 'Units of original result'],
    ['LB-008', 'LBSTRESN', 'Result (Standard)', 'Decimal', 10, 'Auto', '—',
     'Standardized numeric; flag if outside normal range', 'LBSTRESN', 'Standardized numeric result'],
    ['LB-009', 'LBNRIND', 'Normal Range Indicator', 'Coded', 10, 'Auto',
     'NORMAL|LOW|HIGH|CRITICALLY LOW|CRITICALLY HIGH',
     'Based on central lab reference ranges', 'LBNRIND', 'Relationship to reference range'],
    ['LB-010', 'LBSTNRLO', 'Normal Range Lower', 'Decimal', 10, 'Yes', '—',
     'From central lab; test-specific', 'LBSTNRLO', 'Lower limit of normal'],
    ['LB-011', 'LBSTNRHI', 'Normal Range Upper', 'Decimal', 10, 'Yes', '—',
     'From central lab; test-specific', 'LBSTNRHI', 'Upper limit of normal'],
], columns=['Field_ID', 'Variable', 'Label', 'DataType', 'Length',
           'Required', 'CodedList', 'ValidationRule', 'SDTM_Map', 'Description'])

# CM - Concomitant Medications
cm_spec = pd.DataFrame([
    ['CM-001', 'USUBJID', 'Subject Identifier', 'Text', 15, 'Yes', '—',
     'Must exist in DM domain', 'USUBJID', 'Link to demographics'],
    ['CM-002', 'CMSEQ', 'CM Sequence Number', 'Integer', 3, 'Auto', '—',
     'Auto-incremented per subject', 'CMSEQ', 'Unique within subject'],
    ['CM-003', 'CMTRT', 'Medication Name', 'Text', 200, 'Yes', '—',
     'Generic name preferred; coded via WHO Drug', 'CMTRT', 'Verbatim medication name'],
    ['CM-004', 'CMDECOD', 'Standardized Name', 'Text', 200, 'Yes', 'WHO Drug B3',
     'Must be coded if CMTRT present', 'CMDECOD', 'WHO Drug dictionary term'],
    ['CM-005', 'CMDOSE', 'Dose', 'Decimal', 10, 'Yes', '—',
     'Must be > 0', 'CMDOSE', 'Numeric dose per administration'],
    ['CM-006', 'CMDOSU', 'Dose Unit', 'Coded', 10, 'Yes', 'mg|mcg|g|mL|IU',
     'Must be valid unit', 'CMDOSU', 'Unit of dose'],
    ['CM-007', 'CMROUTE', 'Route', 'Coded', 20, 'Yes',
     'ORAL|INTRAVENOUS|SUBCUTANEOUS|TOPICAL|INTRAMUSCULAR|INHALATION',
     'Must be valid route', 'CMROUTE', 'Route of administration'],
    ['CM-008', 'CMSTDTC', 'Start Date', 'Date', 10, 'Yes', '—',
     'Format YYYY-MM-DD; partial dates allowed', 'CMSTDTC', 'Medication start date'],
    ['CM-009', 'CMENDTC', 'End Date', 'Date', 10, 'Cond', '—',
     'Must be ≥ CMSTDTC; blank if ongoing', 'CMENDTC', 'Medication stop date'],
    ['CM-010', 'CMINDC', 'Indication', 'Text', 200, 'Yes', '—',
     'Medical condition being treated', 'CMINDC', 'Reason for taking medication'],
], columns=['Field_ID', 'Variable', 'Label', 'DataType', 'Length',
           'Required', 'CodedList', 'ValidationRule', 'SDTM_Map', 'Description'])

# ============================================================
# WRITE TO EXCEL WITH PROFESSIONAL FORMATTING
# ============================================================
wb = Workbook()

# Styling
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
alt_fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_domain_sheet(wb, name, df, tab_color):
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = tab_color

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

# Remove default sheet and write all domains
wb.remove(wb.active)
write_domain_sheet(wb, 'DM - Demographics', dm_spec, '3498DB')
write_domain_sheet(wb, 'AE - Adverse Events', ae_spec, 'E74C3C')
write_domain_sheet(wb, 'VS - Vital Signs', vs_spec, '2ECC71')
write_domain_sheet(wb, 'LB - Laboratory', lb_spec, 'F39C12')
write_domain_sheet(wb, 'CM - Concomitant Meds', cm_spec, '9B59B6')

# Add cover sheet
cover = wb.create_sheet(title='Cover', index=0)
cover.sheet_properties.tabColor = '2C3E50'
cover_data = [
    ['ELECTRONIC CASE REPORT FORM (eCRF) SPECIFICATION'],
    [''],
    ['Study:', 'XYZ-2026'],
    ['Phase:', 'Phase II, Randomized, Double-Blind, Placebo-Controlled'],
    ['Indication:', 'Type 2 Diabetes Mellitus'],
    ['Sponsor:', 'Simulated Pharma Inc.'],
    ['CRO:', 'Simulated CRO Services'],
    [''],
    ['Document Version:', '1.0'],
    ['Prepared By:', 'Careen Chowrappa, Configuration & QC Specialist'],
    ['Date:', 'March 2026'],
    [''],
    ['DOMAINS INCLUDED:'],
    ['Domain', 'Code', 'Fields', 'Description'],
    ['Demographics', 'DM', '12', 'Subject enrollment and baseline characteristics'],
    ['Adverse Events', 'AE', '12', 'Safety reporting throughout study'],
    ['Vital Signs', 'VS', '10', 'BP, HR, Temp, Weight at each visit'],
    ['Laboratory', 'LB', '11', 'Central lab results (metabolic, hepatic, renal)'],
    ['Concomitant Meds', 'CM', '10', 'Prior and concurrent medications'],
    [''],
    ['TOTAL FIELDS: 55 across 5 SDTM-aligned domains'],
    ['TOTAL EDIT CHECKS: 20 (see Edit Check Specifications sheet)'],
]

for row_idx, row_data in enumerate(cover_data, 1):
    for col_idx, value in enumerate(row_data if isinstance(row_data, list) else [row_data], 1):
        cell = cover.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16, color='2C3E50')
        elif row_idx == 14:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        elif value in ['Study:', 'Phase:', 'Indication:', 'Sponsor:', 'CRO:',
                       'Document Version:', 'Prepared By:', 'Date:']:
            cell.font = Font(bold=True)

cover.column_dimensions['A'].width = 25
cover.column_dimensions['B'].width = 50
cover.column_dimensions['C'].width = 10
cover.column_dimensions['D'].width = 50

wb.save('docs/01_eCRF_configuration_spec.xlsx')
print("\n✅ eCRF Specification saved: docs/01_eCRF_configuration_spec.xlsx")
print(f"   Domains: 5 | Total Fields: 55 | Sheets: 6 (Cover + 5 domains)")
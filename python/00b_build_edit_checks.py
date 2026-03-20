"""
============================================================
00b_build_edit_checks.py
Purpose: Generate Edit Check Specification Document
Study:   XYZ-2026
Author:  Careen Chowrappa
============================================================
Maps to ICON JD: "Validation plans and procedures to assess
the functionality and performance of EDC systems"
============================================================
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

edit_checks = pd.DataFrame([
    # --- DEMOGRAPHIC CHECKS ---
    ['EC-DM-001', 'DM', 'AGE', 'HARD', 'Range',
     'AGE must be between 18 and 75 (per protocol inclusion criteria)',
     'AGE < 18 OR AGE > 75',
     'Subject age [{AGE}] is outside protocol-specified range (18-75). '
     'Please verify against source document.',
     'Protocol Section 5.1 Inclusion Criteria'],
    ['EC-DM-002', 'DM', 'RFSTDTC vs RFICDTC', 'HARD', 'Date Logic',
     'First dose date must be on or after informed consent date',
     'RFSTDTC < RFICDTC',
     'First dose date ({RFSTDTC}) is before informed consent date ({RFICDTC}). '
     'This is a protocol violation. Please review.',
     'ICH E6(R3) Section 4.8'],
    ['EC-DM-003', 'DM', 'SEX', 'HARD', 'Coded Value',
     'Sex must be M or F per controlled terminology',
     'SEX NOT IN ("M", "F")',
     'Invalid sex value "{SEX}". Expected M or F per CDISC controlled terminology.',
     'CDISC CT 2024-03-29'],
    ['EC-DM-004', 'DM', 'AGE vs BRTHDTC', 'HARD', 'Cross-field',
     'Reported age must match age calculated from date of birth and consent date',
     'ABS(AGE - CALCULATED_AGE) > 1',
     'Reported age ({AGE}) does not match calculated age from DOB ({CALC_AGE}). '
     'Discrepancy of {DIFF} years. Please verify.',
     'Data quality — transcription check'],
    ['EC-DM-005', 'DM', 'USUBJID', 'HARD', 'Format',
     'Subject ID must follow format XYZ-2026-SSS-PPPP',
     'NOT REGEX("^XYZ-2026-\\d{3}-\\d{4}$")',
     'Subject ID "{USUBJID}" does not follow expected format XYZ-2026-SSS-PPPP. '
     'Please correct.',
     '21 CFR Part 11 — traceability'],
    ['EC-DM-006', 'DM', 'ALL REQUIRED', 'HARD', 'Completeness',
     'All required DM fields must be non-null',
     'ANY REQUIRED FIELD IS NULL',
     '{FIELD} is missing. This field is required per protocol CRF completion guidelines.',
     'ICH E6(R3) Section 5.5'],

    # --- ADVERSE EVENT CHECKS ---
    ['EC-AE-001', 'AE', 'AESTDTC vs RFICDTC', 'HARD', 'Date Logic',
     'AE start date must be on or after informed consent date',
     'AESTDTC < RFICDTC',
     'AE start date ({AESTDTC}) is before informed consent ({RFICDTC}). '
     'Pre-consent events should not be in AE log unless pre-existing.',
     'ICH E6(R3) Section 6.2'],
    ['EC-AE-002', 'AE', 'AEENDTC vs AESTDTC', 'HARD', 'Date Logic',
     'AE end date must be on or after AE start date',
     'AEENDTC < AESTDTC',
     'AE end date ({AEENDTC}) is before start date ({AESTDTC}). '
     'Temporal impossibility. Please review dates.',
     'Data quality — temporal logic'],
    ['EC-AE-003', 'AE', 'AESEV + AESER', 'SOFT', 'Clinical Logic',
     'If severity is SEVERE, verify if AE should also be marked SERIOUS',
     'AESEV = "SEVERE" AND AESER = "N"',
     'AE is graded SEVERE but not marked as Serious. Per ICH E2A, please '
     'confirm this AE does not meet any SAE criteria.',
     'ICH E2A — SAE definition'],
    ['EC-AE-004', 'AE', 'AEOUT + AEENDTC', 'SOFT', 'Conditional',
     'If outcome is RECOVERED, end date should be provided',
     'AEOUT = "RECOVERED/RESOLVED" AND AEENDTC IS NULL',
     'Outcome is RECOVERED but no end date provided. Please enter resolution date.',
     'Data quality — completeness'],
    ['EC-AE-005', 'AE', 'AETERM + AEDECOD', 'SOFT', 'Coding',
     'If verbatim term is entered, MedDRA coded term must also be populated',
     'AETERM IS NOT NULL AND AEDECOD IS NULL',
     'AE verbatim term "{AETERM}" is entered but MedDRA coded term is missing. '
     'Please complete medical coding.',
     'MedDRA coding requirements'],
    ['EC-AE-006', 'AE', 'AESEV', 'HARD', 'Coded Value',
     'Severity must be MILD, MODERATE, or SEVERE',
     'AESEV NOT IN ("MILD", "MODERATE", "SEVERE")',
     'Invalid severity value "{AESEV}". Expected MILD, MODERATE, or SEVERE.',
     'CDISC CT — AESEV codelist'],

    # --- VITAL SIGNS CHECKS ---
    ['EC-VS-001', 'VS', 'SYSBP', 'HARD', 'Range',
     'Systolic BP must be between 70 and 250 mmHg',
     'SYSBP < 70 OR SYSBP > 250',
     'Systolic BP value {SYSBP} mmHg is outside physiological range (70-250). '
     'Please verify against source.',
     'Clinical plausibility check'],
    ['EC-VS-002', 'VS', 'DIABP', 'HARD', 'Range',
     'Diastolic BP must be between 40 and 150 mmHg',
     'DIABP < 40 OR DIABP > 150',
     'Diastolic BP value {DIABP} mmHg is outside physiological range (40-150). '
     'Please verify.',
     'Clinical plausibility check'],
    ['EC-VS-003', 'VS', 'SYSBP vs DIABP', 'HARD', 'Cross-field',
     'Systolic BP must be greater than Diastolic BP',
     'SYSBP <= DIABP',
     'Systolic ({SYSBP}) ≤ Diastolic ({DIABP}). '
     'Physiologically impossible. Please verify both values.',
     'Clinical plausibility — cross-field'],
    ['EC-VS-004', 'VS', 'PULSE', 'HARD', 'Range',
     'Heart rate must be between 40 and 180 bpm',
     'PULSE < 40 OR PULSE > 180',
     'Heart rate {PULSE} bpm is outside range (40-180). Please verify.',
     'Clinical plausibility check'],
    ['EC-VS-005', 'VS', 'TEMP', 'HARD', 'Range',
     'Temperature must be between 35.0 and 40.0 °C',
     'TEMP < 35.0 OR TEMP > 40.0',
     'Temperature {TEMP}°C is outside expected range (35.0-40.0). '
     'Please verify measurement and units.',
     'Clinical plausibility check'],
    ['EC-VS-006', 'VS', 'ALL REQUIRED', 'HARD', 'Completeness',
     'All protocol-required vital signs must be recorded at each visit',
     'ANY REQUIRED VS FIELD IS NULL',
     '{FIELD} is missing for {VISIT}. All vitals are required per protocol.',
     'Protocol Section 8.2 Assessments'],

    # --- LAB CHECKS ---
    ['EC-LB-001', 'LB', 'HBA1C', 'HARD', 'Range',
     'HbA1c must be between 3.0% and 20.0%',
     'HBA1C < 3.0 OR HBA1C > 20.0',
     'HbA1c value {HBA1C}% is outside testable range (3.0-20.0). '
     'Please verify with central lab.',
     'Clinical plausibility — diabetes-specific'],
    ['EC-LB-002', 'LB', 'ALT', 'SOFT', 'Safety Signal',
     'Flag ALT > 3x upper limit of normal (>120 U/L) as potential hepatotoxicity',
     'ALT > 120',
     'ALT value {ALT} U/L exceeds 3x ULN (>120 U/L). '
     'Potential Hy\'s Law signal. Medical monitor review recommended.',
     'FDA Hy\'s Law Guidance 2009'],
], columns=['Check_ID', 'Domain', 'Field(s)', 'Severity', 'Check_Type',
           'Rule_Description', 'Trigger_Condition', 'Query_Text', 'Regulatory_Reference'])

# Write to Excel with formatting
wb = Workbook()
ws = wb.active
ws.title = 'Edit Check Specifications'
ws.sheet_properties.tabColor = 'E74C3C'

header_font = Font(bold=True, color='FFFFFF', size=10)
header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
hard_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
soft_fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
border = Border(left=Side('thin'), right=Side('thin'),
                top=Side('thin'), bottom=Side('thin'))

for col_idx, col_name in enumerate(edit_checks.columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

for row_idx, row in edit_checks.iterrows():
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if row['Severity'] == 'HARD':
            cell.fill = hard_fill
        else:
            cell.fill = soft_fill

widths = [12, 6, 20, 8, 12, 50, 35, 60, 30]
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65 + i)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:I{len(edit_checks) + 1}'

wb.save('docs/02_edit_check_specs.xlsx')
print("✅ Edit Check Specifications saved: docs/02_edit_check_specs.xlsx")
print(f"   Total checks: {len(edit_checks)} | HARD: "
      f"{len(edit_checks[edit_checks['Severity']=='HARD'])} | "
      f"SOFT: {len(edit_checks[edit_checks['Severity']=='SOFT'])}")